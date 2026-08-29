#!/usr/bin/env python3
from __future__ import annotations

import importlib.util
import json
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import unittest
import zipfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

import manifest_runtime as runtime
import ocr_api


MODULE_PATH = Path(__file__).with_name("run_assessment.py")
SPEC = importlib.util.spec_from_file_location("score_kuaishou_pipeline", MODULE_PATH)
assert SPEC and SPEC.loader
pipeline = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = pipeline
SPEC.loader.exec_module(pipeline)


def document(headers, rows, *, student_name="", source="docs", document_id="test", revision="r1"):
    payload = {
        "schema_version": 1,
        "source": source,
        "document": {"url": f"https://docs.corp.kuaishou.com/{document_id}", "id": document_id, "revision": revision},
        "sheet": "Sheet1",
        "range": "A1:Z100",
        "headers": headers,
        "rows": rows,
        "_source_path": "fixture.json",
    }
    if student_name:
        payload["student_name"] = student_name
    return payload


def write_json(path, payload):
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def manifest_payload(*, standard="standard.json", homework=None, max_concurrency=None):
    runtime_config = {}
    if max_concurrency is not None:
        runtime_config["max_concurrency"] = max_concurrency
    return {
        "schema_version": 1,
        "task_id": "task-29-quiz",
        "exam_profile": "retake-image-aesthetic",
        "group": "29组",
        "progress": "补考画面美学",
        "standard": {
            "id": "standard-doc",
            "url": "https://docs.corp.kuaishou.com/standard-doc",
            "revision": "std-r1",
            "sheet": "Sheet1",
            "range": "A1:Z100",
            "evidence": standard,
        },
        "homework": {
            "documents": homework
            or [
                {
                    "student": "张三",
                    "id": "student-zhang",
                    "url": "https://docs.corp.kuaishou.com/student-zhang",
                    "revision": "zhang-r1",
                    "sheet": "Sheet1",
                    "range": "A1:Z100",
                    "evidence": "zhang.json",
                },
                {
                    "student": "李四",
                    "id": "student-li",
                    "url": "https://docs.corp.kuaishou.com/student-li",
                    "revision": "li-r1",
                    "sheet": "Sheet1",
                    "range": "A1:Z100",
                    "evidence": "li.json",
                },
            ]
        },
        "runtime": runtime_config,
        "output": {"dir": "output", "png": "off", "xlsx": "off"},
    }


class PipelineTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.config = pipeline.load_config()

    def test_profiles_are_fixed_and_complete(self):
        profiles = self.config["profiles"]
        self.assertEqual(len(profiles), 10)
        self.assertEqual(len(profiles["quiz-text-to-video"]["dimensions"]), 6)
        self.assertEqual(len(profiles["final-text-to-video"]["dimensions"]), 10)
        self.assertEqual(len(profiles["final-image-to-video"]["dimensions"]), 11)
        self.assertEqual(profiles["retake-text-to-video"]["dimensions"], profiles["final-text-to-video"]["dimensions"])
        self.assertEqual(profiles["retake-image-to-video"]["dimensions"], profiles["final-image-to-video"]["dimensions"])

    def test_id_normalization(self):
        self.assertEqual(pipeline.normalize_id(" 0046 "), "46")
        self.assertEqual(pipeline.normalize_id("46.0"), "46")
        self.assertEqual(pipeline.normalize_id(46.0), "46")
        self.assertEqual(pipeline.normalize_id("A-046"), "A-046")

    def test_keyword_matching_uses_bidirectional_containment(self):
        cases = [
            (("videoContrast",), "videoContrast好", ("videoContrast",)),
            (("videoExp",), "videoExp好", ("videoExp",)),
            (("都一般", "一样好"), "答案：一样好", ("一样好",)),
            (("都一般", "videoExp好"), "作业选择 videoExp好", ("videoExp好",)),
            (("video Contrast",), "videoContrast好", ("video Contrast",)),
            (("videoContrast好",), "videoContrast", ("videoContrast好",)),
            (("都一般",), "般", ("都一般",)),
        ]
        for keywords, homework, expected in cases:
            with self.subTest(keywords=keywords, homework=homework):
                self.assertEqual(pipeline.match_standard_keywords(homework, keywords), expected)
        self.assertEqual(pipeline.match_standard_keywords("一样差", ("都一般", "一样好")), ())

    def test_standard_multi_answer_text_delimiters(self):
        for separator in ["\n", ",", "，", "、", "/", "／", "|", "｜"]:
            with self.subTest(separator=repr(separator)):
                keywords = pipeline.standard_answer_keywords(f"都一般{separator}一样好")
                self.assertEqual(keywords, ("都一般", "一样好"))
        self.assertEqual(
            pipeline.standard_answer_keywords("都一般，，"),
            ("都一般",),
        )

    def test_standard_multi_answer_structured_values_and_deduplication(self):
        cases = [
            ["videoContrast", "自定义关键词"],
            [{"text": "videoContrast"}, {"value": "自定义关键词"}],
            {"values": [{"label": "videoContrast"}, {"name": "自定义关键词"}]},
        ]
        for raw in cases:
            with self.subTest(raw=raw):
                self.assertEqual(pipeline.standard_answer_keywords(raw), ("videoContrast", "自定义关键词"))
        self.assertEqual(pipeline.standard_answer_keywords(["任意文本", "任意文本", "另一个答案"]), ("任意文本", "另一个答案"))
        self.assertEqual(pipeline.standard_answer_keywords({"value": "internal-option-id", "text": "用户可见关键词"}), ("用户可见关键词",))

    def test_standard_keywords_accept_arbitrary_text_but_reject_mixed_blank(self):
        self.assertEqual(pipeline.standard_answer_keywords("不在旧五值中的原始关键词"), ("不在旧五值中的原始关键词",))
        with self.assertRaisesRegex(pipeline.AssessmentError, "空值与非空关键词混合"):
            pipeline.standard_answer_keywords([None, "一样好"])
        with self.assertRaisesRegex(pipeline.AssessmentError, "结构无法识别"):
            pipeline.standard_answer_keywords({"color": "blue"})

    def test_multi_answer_matches_any_option_and_counts_once(self):
        dimensions = ["画面美学"]
        raw_options = {"values": [{"text": "videoContrast"}, {"text": "自定义关键词"}, {"text": "自定义关键词"}]}
        standard_docs = [document(["ID", "画面美学"], [[1, raw_options], [2, "另一个固定答案"]])]
        homework_docs = [document(["同学名称", "ID", "画面美学"], [["张三", 1, "videoContrast好"], ["张三", 2, "不相关内容"]])]
        standard = pipeline.parse_standard(standard_docs, dimensions, self.config)
        self.assertEqual(standard["画面美学"]["1"]["keywords"], ("videoContrast", "自定义关键词"))
        self.assertEqual(standard["画面美学"]["1"]["raw_value"], raw_options)
        students, anomalies = pipeline.parse_homework_documents(homework_docs, dimensions, self.config, {})
        result = pipeline.score_students(standard, students, dimensions, anomalies)
        dimension = result["summary"][0]["dimensions"]["画面美学"]
        self.assertEqual((dimension["correct"], dimension["total"], dimension["accuracy"]), (1, 2, 0.5))
        first, second = result["details"]
        self.assertEqual(first["standard_keywords"], "videoContrast｜自定义关键词")
        self.assertEqual(first["matched_keywords"], "videoContrast")
        self.assertIn("双向模糊命中标准关键词", first["note"])
        self.assertIn("未与任何标准关键词形成包含关系", second["note"])

    def test_all_profiles_only_read_declared_dimensions(self):
        for profile_key, profile in self.config["profiles"].items():
            dimensions = profile["dimensions"]
            headers = ["ID", *dimensions, "解释", "任意额外维度"]
            row = [46, *["都一般" for _ in dimensions], "不参与", "一样差"]
            parsed = pipeline.parse_standard([document(headers, [row])], dimensions, self.config)
            self.assertEqual(list(parsed), dimensions, profile_key)
            self.assertTrue(all(list(values) == ["46"] for values in parsed.values()), profile_key)

    def test_non_special_standard_blank_requires_decision_before_scoring(self):
        standard_docs = [document(["ID", "画面美学", "动态美学"], [[1, "", ""], [2, "都一般", "一样好"]])]
        homework_docs = [document(["同学名称", "ID", "画面美学", "动态美学"], [["张三", 1, "任意内容", "一样好"], ["张三", 2, "都一般", "一样好"]])]
        result = pipeline.build_scored_result(
            profile_key="online-aesthetics-dual",
            group="29组",
            progress="异常标准空值",
            standard_documents=standard_docs,
            homework_documents=homework_docs,
            aliases={},
            source_mode="docs",
        )
        self.assertEqual(result["schema_version"], 4)
        self.assertEqual(result["run_status"], "awaiting_standard_decision")
        self.assertEqual(result["summary"], [])
        self.assertEqual(result["details"], [])
        self.assertEqual(len(result["evidence"]), 2)
        self.assertEqual(
            {(item["dimension"], item["id"]) for item in result["decision_request"]["affected_cells"]},
            {("画面美学", "1"), ("动态美学", "1")},
        )
        self.assertTrue(result["decision_request"]["data_retained"])
        self.assertTrue(all(item["stage"] == "standard_answer_review" for item in result["stopped_items"]))

    def test_standard_blank_decision_excludes_cells_and_uses_effective_denominator(self):
        standard_docs = [document(["ID", "画面美学", "动态美学"], [[1, "", "一样好"], [2, "都一般", "一样好"]])]
        homework_docs = [document(["同学名称", "ID", "画面美学", "动态美学"], [["张三", 1, "即使非空也排除", "一样好"], ["张三", 2, "都一般", "一样好"]])]
        waiting = pipeline.build_scored_result(
            profile_key="online-aesthetics-dual", group="29组", progress="逐格排除",
            standard_documents=standard_docs, homework_documents=homework_docs,
            aliases={}, source_mode="docs",
        )
        key = waiting["decision_request"]["decision_key"]
        result = pipeline.build_scored_result(
            profile_key="online-aesthetics-dual", group="29组", progress="逐格排除",
            standard_documents=standard_docs, homework_documents=homework_docs,
            aliases={}, source_mode="docs", standard_blank_action="exclude",
            standard_blank_decision_key_value=key,
        )
        self.assertEqual(result["run_status"], "complete")
        summary = result["summary"][0]
        aesthetic = summary["dimensions"]["画面美学"]
        self.assertEqual((aesthetic["source_total"], aesthetic["total"], aesthetic["correct"]), (2, 1, 1))
        self.assertEqual(aesthetic["excluded_ids"], ["1"])
        self.assertEqual((summary["overall_correct"], summary["overall_total"], summary["overall_accuracy"]), (3, 3, 1.0))
        self.assertEqual(result["group_summary"], {"correct": 3, "total": 3, "accuracy": 1.0})
        excluded = [item for item in result["details"] if item["result"] == "不计分"]
        self.assertEqual([(item["dimension"], item["id"]) for item in excluded], [("画面美学", "1")])
        self.assertEqual(result["standard_blank_decision"]["decision_key"], key)
        tampered = json.loads(json.dumps(result, ensure_ascii=False))
        tampered.pop("standard_blank_decision")
        with self.assertRaisesRegex(pipeline.AssessmentError, "standard_blank_decision"):
            pipeline.validate_result_schema(tampered)

        stale = pipeline.build_scored_result(
            profile_key="online-aesthetics-dual", group="29组", progress="逐格排除",
            standard_documents=standard_docs, homework_documents=homework_docs,
            aliases={}, source_mode="docs", standard_blank_action="exclude",
            standard_blank_decision_key_value="stale-key",
        )
        self.assertEqual(stale["run_status"], "awaiting_standard_decision")
        self.assertFalse(stale["decision_request"]["provided_key_valid"])
        self.assertTrue(any("decision_key" in item["reason"] for item in stale["stopped_items"]))
        changed_standard = [document(["ID", "画面美学", "动态美学"], [[1, "都一般", "一样好"], [2, "", "一样好"]])]
        changed = pipeline.build_scored_result(
            profile_key="online-aesthetics-dual", group="29组", progress="逐格排除",
            standard_documents=changed_standard, homework_documents=homework_docs,
            aliases={}, source_mode="docs", standard_blank_action="exclude",
            standard_blank_decision_key_value=key,
        )
        self.assertEqual(changed["run_status"], "awaiting_standard_decision")
        self.assertNotEqual(changed["decision_request"]["decision_key"], key)

        revision_changed_standard = [
            document(
                ["ID", "画面美学", "动态美学"],
                [[1, "", "一样好"], [2, "都一般", "一样好"]],
                revision="r2",
            )
        ]
        revision_changed = pipeline.build_scored_result(
            profile_key="online-aesthetics-dual", group="29组", progress="逐格排除",
            standard_documents=revision_changed_standard, homework_documents=homework_docs,
            aliases={}, source_mode="docs", standard_blank_action="exclude",
            standard_blank_decision_key_value=key,
        )
        self.assertEqual(revision_changed["run_status"], "awaiting_standard_decision")
        self.assertNotEqual(revision_changed["decision_request"]["decision_key"], key)

    def test_standard_blank_gate_reports_evidence_failures_at_the_same_time(self):
        result = pipeline.build_scored_result(
            profile_key="retake-image-aesthetic", group="29组", progress="空值与证据失败",
            standard_documents=[document(["ID", "画面美学"], [[1, ""]])],
            homework_documents=[document(["同学名称", "ID", "画面美学"], [["张三", 1, "任意"]])],
            aliases={}, source_mode="docs",
            failed_documents=[{"role": "homework", "student": "李四", "error": "无权访问"}],
        )
        self.assertEqual(result["run_status"], "awaiting_standard_decision")
        self.assertEqual({item["stage"] for item in result["stopped_items"]}, {"standard_answer_review", "evidence"})
        self.assertTrue(all(item["reason"] and item["next_action"] for item in result["stopped_items"]))

    def test_standard_blank_gate_reports_ocr_review_and_evidence_failure_together(self):
        standard_doc = document(["ID", "画面美学"], [[1, ""]], source="image_ocr")
        standard_doc["confidence"] = {"1": 0.4}
        standard_doc["confidence_issues"] = {}
        homework_doc = document(
            ["同学名称", "ID", "画面美学"],
            [["张三", 1, "任意"]],
            source="image_ocr",
        )
        homework_doc["confidence"] = {"1": 0.3}
        homework_doc["confidence_issues"] = {}
        result = pipeline.build_scored_result(
            profile_key="retake-image-aesthetic", group="29组", progress="组合停止原因",
            standard_documents=[standard_doc], homework_documents=[homework_doc],
            aliases={}, source_mode="images", ocr_confidence_threshold=0.75,
            failed_documents=[{"role": "homework", "student": "李四", "error": "无权访问"}],
        )
        self.assertEqual(result["run_status"], "awaiting_standard_decision")
        self.assertEqual(result["details"], [])
        self.assertEqual(
            {item["stage"] for item in result["stopped_items"]},
            {"standard_answer_review", "ocr_review", "evidence"},
        )
        ocr_stops = [item for item in result["stopped_items"] if item["stage"] == "ocr_review"]
        self.assertEqual({item["role"] for item in ocr_stops}, {"standard", "homework"})
        self.assertTrue(all(item["reason"] and item["next_action"] for item in ocr_stops))

    def test_fully_excluded_dimension_is_unscored_but_other_dimensions_count(self):
        standard_docs = [document(["ID", "画面美学", "动态美学"], [[1, "", "一样好"], [2, "", "一样好"]])]
        homework_docs = [document(["同学名称", "ID", "画面美学", "动态美学"], [["张三", 1, "任意", "一样好"], ["张三", 2, "", "错误"]])]
        waiting = pipeline.build_scored_result(
            profile_key="online-aesthetics-dual", group="29组", progress="整维未评分",
            standard_documents=standard_docs, homework_documents=homework_docs,
            aliases={}, source_mode="docs",
        )
        result = pipeline.build_scored_result(
            profile_key="online-aesthetics-dual", group="29组", progress="整维未评分",
            standard_documents=standard_docs, homework_documents=homework_docs,
            aliases={}, source_mode="docs", standard_blank_action="exclude",
            standard_blank_decision_key_value=waiting["decision_request"]["decision_key"],
        )
        summary = result["summary"][0]
        self.assertEqual(summary["dimensions"]["画面美学"]["status"], "未评分")
        self.assertIsNone(summary["dimensions"]["画面美学"]["accuracy"])
        self.assertEqual((summary["overall_correct"], summary["overall_total"], summary["overall_accuracy"]), (1, 2, 0.5))
        self.assertEqual(result["run_status"], "complete")

    def test_all_cells_excluded_stops_formal_delivery(self):
        standard_docs = [document(["ID", "画面美学"], [[1, ""], [2, ""]])]
        homework_docs = [document(["同学名称", "ID", "画面美学"], [["张三", 1, "任意"], ["张三", 2, ""]])]
        waiting = pipeline.build_scored_result(
            profile_key="retake-image-aesthetic", group="29组", progress="无有效格",
            standard_documents=standard_docs, homework_documents=homework_docs,
            aliases={}, source_mode="docs",
        )
        result = pipeline.build_scored_result(
            profile_key="retake-image-aesthetic", group="29组", progress="无有效格",
            standard_documents=standard_docs, homework_documents=homework_docs,
            aliases={}, source_mode="docs", standard_blank_action="exclude",
            standard_blank_decision_key_value=waiting["decision_request"]["decision_key"],
        )
        self.assertEqual(result["run_status"], "incomplete")
        self.assertEqual(result["summary"][0]["status"], "未评分")
        self.assertEqual(result["group_summary"], {"correct": 0, "total": 0, "accuracy": None})
        self.assertIn("所有维度均无有效评分格", result["failed_documents"][0]["error"])

    def test_special_blank_matrix(self):
        dimensions = ["多镜头指令遵循", "多镜头间一致连贯性"]
        standard_docs = [document(["ID", *dimensions], [[1, "", ""], [2, "videoExp", "videoContrast"]])]
        homework_docs = [
            document(
                ["同学名称", "ID", *dimensions],
                [
                    ["张三", 2, "videoExp好", ""],
                    ["张三", 1, "", "C好"],
                ],
            )
        ]
        standard = pipeline.parse_standard(standard_docs, dimensions, self.config)
        self.assertEqual(pipeline.unexpected_standard_blanks(standard, dimensions, self.config), [])
        students, anomalies = pipeline.parse_homework_documents(homework_docs, dimensions, self.config, {})
        result = pipeline.score_students(standard, students, dimensions, anomalies)
        first, second = [result["summary"][0]["dimensions"][dim] for dim in dimensions]
        self.assertEqual((first["correct"], first["total"]), (2, 2))
        self.assertEqual((second["correct"], second["total"]), (0, 2))
        self.assertEqual(second["wrong_ids"], ["1", "2"])

    def test_shuffled_ids_missing_extra_and_equal_weight(self):
        dimensions = ["画面美学", "动态美学"]
        standard_docs = [document(["ID", *dimensions], [[1, "都一般", "一样好"], [2, "videoExp", "videoContrast"]])]
        homework_docs = [
            document(
                ["同学名称", "动态美学", "ID", "画面美学", "解释"],
                [
                    ["李四", "videoContrast好", 2, "videoExp好", "忽略"],
                    ["李四", "一样差", 3, "般", "额外ID"],
                    ["李四", "一样好", 1, "错误文本", "待复核"],
                ],
            )
        ]
        standard = pipeline.parse_standard(standard_docs, dimensions, self.config)
        students, anomalies = pipeline.parse_homework_documents(homework_docs, dimensions, self.config, {})
        result = pipeline.score_students(standard, students, dimensions, anomalies)
        summary = result["summary"][0]
        self.assertEqual(summary["dimensions"]["画面美学"]["accuracy"], 0.5)
        self.assertEqual(summary["dimensions"]["动态美学"]["accuracy"], 1.0)
        self.assertEqual(summary["overall_accuracy"], 0.75)
        self.assertTrue(any(item["type"] == "额外ID" and item["id"] == "3" for item in result["anomalies"]))

    def test_duplicate_id_is_fatal(self):
        with self.assertRaisesRegex(pipeline.AssessmentError, "重复 ID"):
            pipeline.parse_standard(
                [document(["ID", "画面美学"], [[46, "都一般"], [46, "一样好"]])],
                ["画面美学"],
                self.config,
            )

    def test_evidence_schema_version_is_required(self):
        with tempfile.TemporaryDirectory() as temp_name:
            path = Path(temp_name) / "bad.json"
            payload = document(["ID", "画面美学"], [[46, "都一般"]])
            payload["schema_version"] = 2
            path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
            with self.assertRaisesRegex(pipeline.AssessmentError, "schema_version"):
                pipeline.load_evidence_files([path])

    def test_synthetic_ocr_handles_merged_and_split_cells(self):
        def hit(text, x, y, width=0.05, confidence=1.0):
            return {"text": text, "x": x, "y": y, "width": width, "height": 0.03, "confidence": confidence}

        hits = [
            hit("ID", 0.02, 0.94, 0.02),
            hit("动态美学", 0.10, 0.94, 0.07),
            hit("videoContrast动态美学归因", 0.28, 0.94, 0.20),
            hit("46 videoContrast好", 0.08, 0.84, 0.14, 0.5),
            hit("54", 0.08, 0.74, 0.02),
            hit("都一般", 0.11, 0.74, 0.05),
            hit("3", 0.003, 0.74, 0.01),
        ]
        rows, confidence = pipeline.parse_ocr_rows(hits, "动态美学")
        self.assertEqual(rows, [["46", "videoContrast好"], ["54", "都一般"]])
        self.assertEqual(confidence["46"], 0.5)

    def test_filename_name_detection_and_ocr_name_fallback(self):
        aliases = {"小王": "王五"}
        self.assertEqual(pipeline.filename_student(Path("王五.png"), aliases, set()), "王五")
        self.assertEqual(pipeline.filename_student(Path("小王.png"), aliases, {"王五"}), "王五")
        self.assertEqual(pipeline.filename_student(Path("截图_001.png"), aliases, set()), "")
        hits = [
            {"text": "学员姓名", "x": 0.1, "y": 0.9, "width": 0.08, "height": 0.03},
            {"text": "王五", "x": 0.2, "y": 0.9, "width": 0.05, "height": 0.03},
        ]
        self.assertEqual(pipeline.student_name_from_hits(hits, aliases), "王五")

    def test_screenshot_standard_and_homework_use_same_header_id_scoring(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            standard_image = temp / "标准答案.png"
            homework_dir = temp / "homework"
            homework_dir.mkdir()
            homework_image = homework_dir / "截图_001.png"
            standard_image.write_bytes(b"standard")
            homework_image.write_bytes(b"homework")

            def fake_api(paths, *, role, dimensions, workers):
                rows = [["46", "videoContrast"]] if role == "standard" else [["46", "videoContrast好"]]
                return [
                    {
                        "path": str(path.resolve()),
                        "headers": ["题目ID（order）", "画面美学答案"],
                        "rows": rows,
                        "student_name": "" if role == "standard" else "王五",
                        "confidence": {"46": 0.99},
                        "engine": "api",
                        "elapsed_seconds": 0.1,
                        "batch_elapsed_seconds": 0.1,
                        "workers": workers,
                    }
                    for path in paths
                ]

            args = pipeline.parse_args([
                "--exam-profile", "screenshot-image-aesthetic",
                "--group", "29组",
                "--progress", "截图考试",
                "--standard-images", str(standard_image),
                "--images", str(homework_dir),
                "--ocr-engine", "api",
                "--ocr-workers", "4",
                "--output", str(temp / "delivery"),
            ])
            with patch.object(pipeline, "run_api_ocr", side_effect=fake_api):
                _, _, result = pipeline.run(args)
            self.assertEqual(result["run_status"], "complete")
            self.assertEqual(result["summary"][0]["student"], "王五")
            self.assertEqual(result["summary"][0]["overall_accuracy"], 1.0)
            self.assertEqual(result["metadata"]["ocr"]["homework"]["workers"], 4)

    def test_screenshot_manifest_honors_image_source_and_runtime_ocr_settings(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            standard_image = temp / "standard.png"
            homework_image = temp / "截图001.png"
            standard_image.write_bytes(b"standard")
            homework_image.write_bytes(b"homework")
            manifest = {
                "schema_version": 1, "task_id": "screenshot-task", "exam_profile": "screenshot-image-aesthetic",
                "group": "29组", "progress": "截图Manifest",
                "standard": {"images": str(standard_image)},
                "homework": {"images": str(homework_image)},
                "output": {"dir": str(temp / "delivery"), "png": "off", "xlsx": "off"},
                "runtime": {"ocr_engine": "api", "ocr_workers": 3, "ocr_confidence_threshold": 0.8},
            }
            manifest_path = temp / "task.json"
            write_json(manifest_path, manifest)

            def fake_api(paths, *, role, dimensions, workers):
                return [{
                    "path": str(paths[0].resolve()), "headers": ["ID", "画面美学"],
                    "rows": [[1, "都一般"]], "student_name": "" if role == "standard" else "王五",
                    "confidence": {"1": 0.95}, "elapsed_seconds": 0.1,
                    "batch_elapsed_seconds": 0.1, "workers": workers,
                }]

            args = pipeline.parse_args(["--manifest", str(manifest_path)])
            with patch.object(pipeline, "run_api_ocr", side_effect=fake_api):
                _, _, result = pipeline.run(args)
            self.assertEqual(result["run_status"], "complete")
            self.assertEqual(result["metadata"]["source_mode"], "images")
            self.assertEqual(result["metadata"]["ocr"]["homework"]["workers"], 3)
            self.assertEqual(result["metadata"]["ocr_confidence_threshold"], 0.8)

    def test_low_ocr_confidence_is_pending_not_silently_scored(self):
        dimensions = ["画面美学"]
        standard = pipeline.parse_standard([document(["ID", "画面美学"], [[1, "都一般"]])], dimensions, self.config)
        homework = document(["ID", "画面美学"], [[1, "都一般"]], student_name="王五", source="image_ocr")
        homework["confidence"] = {"1": 0.4}
        students, anomalies = pipeline.parse_homework_documents([homework], dimensions, self.config, {})
        result = pipeline.score_students(standard, students, dimensions, anomalies, 0.75)
        self.assertEqual(result["summary"][0]["status"], "待复核")
        self.assertEqual(result["summary"][0]["dimensions"]["画面美学"]["review_ids"], ["1"])

    def test_pending_review_blocks_formal_outputs_and_reports_every_stop_reason(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            standard = document(["ID", "画面美学"], [[1, "都一般"]], source="image_ocr")
            standard["confidence"] = {"1": 0.99}
            homework = document(["ID", "画面美学"], [[1, "都一般"]], student_name="王五", source="image_ocr")
            homework["confidence"] = {"1": 0.4}
            standard_path, homework_path = temp / "standard.json", temp / "homework.json"
            write_json(standard_path, standard)
            write_json(homework_path, homework)
            args = pipeline.parse_args([
                "--exam-profile", "screenshot-image-aesthetic", "--group", "29组", "--progress", "待复核门禁",
                "--standard-evidence", str(standard_path), "--homework-evidence", str(homework_path),
                "--output", str(temp / "delivery"), "--png", "on", "--xlsx", "on",
            ])
            xlsx, png, result = pipeline.run(args)
            self.assertEqual(result["run_status"], "pending_review")
            self.assertIsNone(xlsx)
            self.assertIsNone(png)
            self.assertEqual(result["outputs"]["xlsx"], None)
            self.assertEqual(result["outputs"]["png"], None)
            self.assertTrue(result["stopped_items"])
            self.assertTrue(all(item["reason"] and item["next_action"] for item in result["stopped_items"]))
            self.assertTrue(Path(result["_result_path"]).exists())
            completed = subprocess.run(
                [
                    sys.executable, str(MODULE_PATH),
                    "--exam-profile", "screenshot-image-aesthetic", "--group", "29组", "--progress", "待复核门禁CLI",
                    "--standard-evidence", str(standard_path), "--homework-evidence", str(homework_path),
                    "--output", str(temp / "cli-delivery"), "--png", "on", "--xlsx", "on",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(completed.returncode, 5, completed.stderr)
            cli_payload = json.loads(completed.stdout.strip().splitlines()[-1])
            self.assertEqual(cli_payload["status"], "pending_review")
            self.assertTrue(cli_payload["stopped_items"][0]["reason"])

    def test_standard_blank_gate_cli_exits_six_and_never_emits_formal_outputs(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            standard_path = temp / "standard.json"
            homework_path = temp / "homework.json"
            write_json(standard_path, document(["ID", "画面美学"], [[1, ""], [2, "都一般"]]))
            write_json(homework_path, document(["同学名称", "ID", "画面美学"], [["张三", 1, "非空"], ["张三", 2, "都一般"]]))
            completed = subprocess.run(
                [
                    sys.executable, str(MODULE_PATH),
                    "--exam-profile", "retake-image-aesthetic", "--group", "29组", "--progress", "标准空值门禁",
                    "--standard-evidence", str(standard_path), "--homework-evidence", str(homework_path),
                    "--output", str(temp / "delivery"), "--png", "on", "--xlsx", "on",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(completed.returncode, 6, completed.stderr)
            payload = json.loads(completed.stdout.strip().splitlines()[-1])
            self.assertEqual(payload["status"], "awaiting_standard_decision")
            self.assertIsNone(payload["xlsx"])
            self.assertIsNone(payload["png"])
            self.assertTrue(payload["decision_request"]["decision_key"])
            self.assertEqual(payload["decision_request"]["affected_cells"][0]["dimension"], "画面美学")
            saved = json.loads(Path(payload["json"]).read_text(encoding="utf-8"))
            self.assertEqual(saved["summary"], [])
            self.assertEqual(len(saved["evidence"]), 2)
            self.assertTrue(saved["decision_request"]["data_retained"])
            self.assertFalse(list((temp / "delivery").rglob("*.xlsx")))
            self.assertFalse(list((temp / "delivery").rglob("*.png")))
            rerender = subprocess.run(
                [
                    sys.executable, str(MODULE_PATH), "--result-json", payload["json"],
                    "--output", str(temp / "rerender"), "--png", "on", "--xlsx", "on",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(rerender.returncode, 6, rerender.stderr)
            rerender_payload = json.loads(rerender.stdout.strip().splitlines()[-1])
            self.assertIsNone(rerender_payload["xlsx"])
            self.assertIsNone(rerender_payload["png"])
            continued = subprocess.run(
                [
                    sys.executable, str(MODULE_PATH),
                    "--exam-profile", "retake-image-aesthetic", "--group", "29组", "--progress", "标准空值门禁",
                    "--standard-evidence", str(standard_path), "--homework-evidence", str(homework_path),
                    "--standard-blank-action", "exclude",
                    "--standard-blank-decision-key", payload["decision_request"]["decision_key"],
                    "--output", str(temp / "continued"), "--png", "on", "--xlsx", "on",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(continued.returncode, 0, continued.stderr)
            continued_payload = json.loads(continued.stdout.strip().splitlines()[-1])
            self.assertEqual(continued_payload["status"], "complete")
            self.assertTrue(Path(continued_payload["xlsx"]).exists())
            self.assertTrue(Path(continued_payload["png"]).exists())

    def test_unscored_dimension_renders_as_unscored_in_excel_and_png(self):
        with tempfile.TemporaryDirectory() as temp_name:
            standard_docs = [document(["ID", "画面美学", "动态美学"], [[1, "", "一样好"], [2, "", "一样好"]])]
            homework_docs = [document(["同学名称", "ID", "画面美学", "动态美学"], [["张三", 1, "任意", "一样好"], ["张三", 2, "", "错误"]])]
            waiting = pipeline.build_scored_result(
                profile_key="online-aesthetics-dual", group="29组", progress="未评分报表",
                standard_documents=standard_docs, homework_documents=homework_docs,
                aliases={}, source_mode="docs",
            )
            result = pipeline.build_scored_result(
                profile_key="online-aesthetics-dual", group="29组", progress="未评分报表",
                standard_documents=standard_docs, homework_documents=homework_docs,
                aliases={}, source_mode="docs", standard_blank_action="exclude",
                standard_blank_decision_key_value=waiting["decision_request"]["decision_key"],
            )
            xlsx, png, _, _, _ = pipeline.deliver_result(result, Path(temp_name) / "delivery", "on", "on")
            self.assertTrue(xlsx and xlsx.exists())
            self.assertTrue(png and png.exists())
            workbook = load_workbook(xlsx, data_only=False)
            values = [cell.value for cell in workbook["成绩汇总"][2]]
            workbook.close()
            self.assertEqual(values[3], "未评分")
            self.assertEqual(values[-1], 0.5)

    def test_standard_and_missing_or_misaligned_homework_confidence_require_review(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)

            def loaded(std_conf, homework_conf):
                std = document(["ID", "画面美学"], [[1, "都一般"]], source="image_ocr")
                hw = document(["ID", "画面美学"], [[1, "都一般"]], student_name="王五", source="image_ocr")
                if std_conf is not None:
                    std["confidence"] = std_conf
                if homework_conf is not None:
                    hw["confidence"] = homework_conf
                std_path, hw_path = temp / "std.json", temp / "hw.json"
                write_json(std_path, std)
                write_json(hw_path, hw)
                return pipeline.load_evidence_files([std_path]), pipeline.load_evidence_files([hw_path])

            for label, std_conf, homework_conf in [
                ("low-standard", {"1": 0.1}, {"1": 0.99}),
                ("missing-homework", {"1": 0.99}, None),
                ("normalized-low-key", {"1": 0.99}, {"1.0": 0.1}),
                ("invalid-homework", {"1": 0.99}, {"1": "0.1"}),
            ]:
                with self.subTest(label=label):
                    standard_docs, homework_docs = loaded(std_conf, homework_conf)
                    result = pipeline.build_scored_result(
                        profile_key="screenshot-image-aesthetic", group="29组", progress=label,
                        standard_documents=standard_docs, homework_documents=homework_docs,
                        aliases={}, source_mode="images", ocr_confidence_threshold=0.75,
                    )
                    self.assertEqual(result["run_status"], "pending_review")
                    self.assertEqual(result["summary"][0]["status"], "待复核")
                    self.assertTrue(result["stopped_items"][0]["reason"])

    def test_score_cache_invalidates_when_ocr_confidence_changes(self):
        with tempfile.TemporaryDirectory() as temp_name:
            config = self.config
            dimensions = ["画面美学"]
            student_doc = document(["ID", "画面美学"], [[1, "都一般"]], student_name="王五", source="image_ocr")
            student_doc["confidence"] = {"1": 0.99}
            students, anomalies = pipeline.parse_homework_documents([student_doc], dimensions, config, {})

            def parsed_standard(confidence):
                standard_doc = document(["ID", "画面美学"], [[1, "都一般"]], source="image_ocr")
                standard_doc["confidence"] = {"1": confidence}
                return pipeline.parse_standard([standard_doc], dimensions, config)

            profile = config["profiles"]["screenshot-image-aesthetic"]
            high, high_stats, _ = pipeline.score_students_incremental(
                parsed_standard(0.99), students, dimensions, anomalies, profile, {}, Path(temp_name), 0.75,
            )
            low, low_stats, _ = pipeline.score_students_incremental(
                parsed_standard(0.1), students, dimensions, anomalies, profile, {}, Path(temp_name), 0.75,
            )
            self.assertEqual(high["summary"][0]["status"], "已完成")
            self.assertEqual(low["summary"][0]["status"], "待复核")
            self.assertEqual(high_stats["misses"], 1)
            self.assertEqual(low_stats["misses"], 1)

    def test_score_cache_key_includes_standard_blank_policy_and_decision_key(self):
        with tempfile.TemporaryDirectory() as temp_name:
            dimensions = ["画面美学"]
            standard = pipeline.parse_standard(
                [document(["ID", "画面美学"], [[1, ""], [2, "都一般"]])],
                dimensions,
                self.config,
            )
            students, anomalies = pipeline.parse_homework_documents(
                [document(["ID", "画面美学"], [[1, "非空"], [2, "都一般"]], student_name="王五")],
                dimensions,
                self.config,
                {},
            )
            profile = self.config["profiles"]["retake-image-aesthetic"]
            first, first_stats, _ = pipeline.score_students_incremental(
                standard, students, dimensions, anomalies, profile, {}, Path(temp_name), 0.75,
                {("画面美学", "1")}, "exclude", "decision-a",
            )
            second, second_stats, _ = pipeline.score_students_incremental(
                standard, students, dimensions, anomalies, profile, {}, Path(temp_name), 0.75,
                {("画面美学", "1")}, "exclude", "decision-a",
            )
            changed, changed_stats, _ = pipeline.score_students_incremental(
                standard, students, dimensions, anomalies, profile, {}, Path(temp_name), 0.75,
                {("画面美学", "1")}, "exclude", "decision-b",
            )
            self.assertEqual(first_stats, {"hits": 0, "misses": 1})
            self.assertEqual(second_stats, {"hits": 1, "misses": 0})
            self.assertEqual(changed_stats, {"hits": 0, "misses": 1})
            self.assertEqual(first["summary"], second["summary"])
            self.assertEqual(first["summary"], changed["summary"])

    def test_ocr_api_confidence_contract_is_strict_and_normalizes_ids(self):
        base = {"student_name": "王五", "headers": ["ID", "画面美学"], "rows": [["46.0", "都一般"]]}
        valid = ocr_api._parse_json_text(json.dumps({**base, "confidence": {"46.0": 0.9}}, ensure_ascii=False))
        self.assertEqual(valid["confidence"], {"46": 0.9})
        for confidence in (None, {"46.0": "0.1"}, {"46.0": True}, {"46.0": 2.5}, {"99": 0.9}):
            payload = dict(base)
            if confidence is not None:
                payload["confidence"] = confidence
            with self.subTest(confidence=confidence), self.assertRaises(ocr_api.ApiOcrError):
                ocr_api._parse_json_text(json.dumps(payload, ensure_ascii=False))

    def test_image_profile_rejects_docs_homework_evidence(self):
        standard = document(["ID", "画面美学"], [[1, "都一般"]], source="image_ocr")
        standard["confidence"] = {"1": 0.99}
        homework = document(["ID", "画面美学"], [[1, "都一般"]], student_name="王五", source="docs")
        with self.assertRaisesRegex(pipeline.AssessmentError, "作业必须来自 image_ocr"):
            pipeline.build_scored_result(
                profile_key="screenshot-image-aesthetic", group="29组", progress="来源门禁",
                standard_documents=[standard], homework_documents=[homework], aliases={}, source_mode="images",
            )

    def test_api_ocr_processes_screenshot_homework_concurrently(self):
        with tempfile.TemporaryDirectory() as temp_name:
            paths = []
            for index in range(4):
                path = Path(temp_name) / f"image-{index}.png"
                path.write_bytes(b"test")
                paths.append(path)
            lock = threading.Lock()
            active = 0
            maximum_active = 0

            def fake_recognize(path, *, role, dimensions):
                nonlocal active, maximum_active
                with lock:
                    active += 1
                    maximum_active = max(maximum_active, active)
                time.sleep(0.03)
                with lock:
                    active -= 1
                return {
                    "path": str(path.resolve()), "engine": "api", "student_name": "王五",
                    "headers": ["ID", "画面美学"], "rows": [["1", "都一般"]],
                    "confidence": {"1": 0.9}, "elapsed_seconds": 0.03,
                }

            with patch.object(ocr_api, "recognize_one", side_effect=fake_recognize):
                results = ocr_api.recognize_many(
                    paths, role="homework", dimensions=["画面美学"], workers=4,
                )
            self.assertEqual([Path(item["path"]).name for item in results], [path.name for path in paths])
            self.assertGreater(maximum_active, 1)
            self.assertTrue(all(item["workers"] == 4 for item in results))

    def test_manifest_defaults_to_15_concurrency_and_rejects_more(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            path = temp / "task.json"
            write_json(path, manifest_payload())
            loaded = runtime.load_manifest(path)
            self.assertEqual(loaded["runtime"]["max_concurrency"], 15)
            payload = manifest_payload(max_concurrency=16)
            write_json(path, payload)
            with self.assertRaisesRegex(runtime.ManifestError, "1–15"):
                runtime.load_manifest(path)

    def test_empty_url_does_not_create_fake_document_id(self):
        with self.assertRaisesRegex(runtime.ManifestError, "url 或 id"):
            runtime.normalize_document_spec({}, role="homework", student="王五")

    def test_preflight_snapshot_must_be_complete_unique_and_known(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            manifest_path = temp / "task.json"
            write_json(manifest_path, manifest_payload())
            manifest = runtime.load_manifest(manifest_path)
            specs = runtime.document_specs(manifest, runtime.load_state(manifest), stage="students")
            snapshot = temp / "snapshot.json"
            write_json(snapshot, {"items": [{"item_id": specs[0]["item_id"], "revision": "r1"}]})
            with self.assertRaisesRegex(runtime.ManifestError, "缺失"):
                runtime.merge_revisions(specs, snapshot)
            write_json(snapshot, {"items": [
                {"item_id": specs[0]["item_id"]},
                {"item_id": specs[0]["item_id"]},
            ]})
            with self.assertRaisesRegex(runtime.ManifestError, "重复"):
                runtime.merge_revisions(specs, snapshot)

    def test_evidence_row_name_must_match_planned_student(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            manifest_path = temp / "task.json"
            write_json(manifest_path, manifest_payload())
            manifest = runtime.load_manifest(manifest_path)
            plan = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="students"))
            zhang = next(item for item in plan["read"] if item["student"] == "张三")
            bad = temp / "bad.json"
            write_json(bad, document(
                ["同学名称", "ID", "画面美学"],
                [["李四", 1, "都一般"]],
                document_id=zhang["document_id"],
                revision=zhang["revision"],
            ))
            with self.assertRaisesRegex(runtime.ManifestError, "行内姓名"):
                runtime.ingest_evidence(manifest, zhang["item_id"], bad)

    def test_concurrent_ingest_preserves_both_students(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            manifest_path = temp / "task.json"
            write_json(manifest_path, manifest_payload())
            manifest = runtime.load_manifest(manifest_path)
            plan = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="students"))
            files = []
            for item in plan["read"]:
                path = temp / f"{item['student']}.json"
                write_json(path, document(
                    ["ID", "画面美学"], [[1, "都一般"]], student_name=item["student"],
                    document_id=item["document_id"], revision=item["revision"],
                ))
                files.append((item, path))
            with ThreadPoolExecutor(max_workers=2) as executor:
                results = list(executor.map(lambda pair: runtime.ingest_evidence(manifest, pair[0]["item_id"], pair[1]), files))
            self.assertEqual({item["status"] for item in results}, {"success"})
            state = runtime.load_state(manifest)
            self.assertTrue(all(state["documents"][item["item_id"]]["status"] == "success" for item, _ in files))

    def test_old_lock_owner_cannot_delete_replacement_lock(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            manifest_path = temp / "task.json"
            write_json(manifest_path, manifest_payload())
            manifest = runtime.load_manifest(manifest_path)
            lock_path = runtime.state_path(manifest).with_suffix(".lock")
            with runtime.state_lock(manifest):
                replacement = {"pid": 99999999, "token": "replacement-owner", "created_at": time.time()}
                lock_path.write_text(json.dumps(replacement), encoding="utf-8")
            self.assertTrue(lock_path.exists())
            self.assertEqual(json.loads(lock_path.read_text(encoding="utf-8"))["token"], "replacement-owner")

    def test_specs_command_lists_stable_item_ids_without_writing_state(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            manifest_path = temp / "task.json"
            write_json(manifest_path, manifest_payload())
            loaded = runtime.load_manifest(manifest_path)
            self.assertFalse(runtime.state_path(loaded).exists())
            completed = subprocess.run(
                [sys.executable, str(Path(runtime.__file__)), "specs", "--manifest", str(manifest_path), "--stage", "students"],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(completed.returncode, 0, completed.stderr)
            payload = json.loads(completed.stdout)
            self.assertEqual([item["student"] for item in payload["items"]], ["张三", "李四"])
            self.assertTrue(all(item["item_id"] for item in payload["items"]))
            self.assertFalse(runtime.state_path(loaded).exists())

    def test_revision_cache_and_failed_document_resume(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            path = temp / "task.json"
            write_json(path, manifest_payload())
            manifest = runtime.load_manifest(path)

            standard_spec = runtime.document_specs(manifest, runtime.load_state(manifest), stage="initial")
            initial = runtime.plan_reads(manifest, standard_spec)
            self.assertEqual(initial["max_concurrency"], 15)
            self.assertEqual(len(initial["read"]), 1)
            standard_file = temp / "standard-evidence.json"
            write_json(standard_file, document(["ID", "画面美学"], [[1, "都一般"]], document_id="standard-doc", revision="std-r1"))
            standard_item = initial["read"][0]["item_id"]
            runtime.ingest_evidence(manifest, standard_item, standard_file)
            repeated = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="initial"))
            self.assertEqual(len(repeated["cached"]), 1)
            self.assertEqual(repeated["read"], [])

            student_specs = runtime.document_specs(manifest, runtime.load_state(manifest), stage="students")
            students_plan = runtime.plan_reads(manifest, student_specs)
            by_student = {item["student"]: item for item in students_plan["read"]}
            zhang_file = temp / "zhang-evidence.json"
            write_json(zhang_file, document(["ID", "画面美学"], [[1, "都一般"]], student_name="张三", document_id="student-zhang", revision="zhang-r1"))
            runtime.ingest_evidence(manifest, by_student["张三"]["item_id"], zhang_file)
            runtime.record_failure(manifest, by_student["李四"]["item_id"], "Docs timeout", "timeout")
            retry_state = runtime.load_state(manifest)
            retry_state["documents"][by_student["李四"]["item_id"]]["next_attempt_at"] = "2000-01-01T00:00:00+00:00"
            runtime.save_state(manifest, retry_state)
            resume = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="students"))
            self.assertEqual([item["student"] for item in resume["cached"]], ["张三"])
            self.assertEqual([item["student"] for item in resume["read"]], ["李四"])

    def test_first_homework_learns_range_and_reuses_it_for_same_task(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            homework = [
                {"student": "张三", "id": "student-zhang", "url": "https://docs.corp.kuaishou.com/student-zhang", "revision": "zhang-r1"},
                {"student": "李四", "id": "student-li", "url": "https://docs.corp.kuaishou.com/student-li", "revision": "li-r1"},
            ]
            payload = manifest_payload(homework=homework)
            payload["homework"]["layout_reuse"] = {"enabled": True, "discovery_range": "A1:AB200"}
            manifest_path = temp / "task.json"
            write_json(manifest_path, payload)
            manifest = runtime.load_manifest(manifest_path)

            student_specs = runtime.document_specs(manifest, runtime.load_state(manifest), stage="students")
            preflight = temp / "preflight.json"
            write_json(preflight, {"items": [
                {"item_id": item["item_id"], "revision": item["revision"], "sheet": "作业", "range": "A1:AB200"}
                for item in student_specs
            ]})
            first_plan = runtime.plan_reads(manifest, runtime.merge_revisions(student_specs, preflight))
            self.assertEqual(len(first_plan["read"]), 1)
            self.assertEqual(first_plan["read"][0]["read_mode"], "discovery_probe")
            self.assertEqual(first_plan["read"][0]["discovery_range"], "A1:AB200")
            self.assertEqual(first_plan["read"][0]["range"], "")
            self.assertEqual([item["student"] for item in first_plan["deferred"]], ["李四"])
            self.assertEqual({item["student"] for item in pipeline.manifest_failures(manifest)}, {"张三", "李四"})

            probe = first_plan["read"][0]
            probe_payload = document(["order", "画面美学"], [[1, "都一般"], [2, "一样好"]], student_name="张三", document_id="student-zhang", revision="zhang-r1")
            probe_payload["sheet"] = "作业"
            probe_payload["range"] = "A1:AB3"
            probe_file = temp / "probe.json"
            write_json(probe_file, probe_payload)
            ingested = runtime.ingest_evidence(manifest, probe["item_id"], probe_file)
            self.assertEqual(ingested["learned_layout"]["range"], "A1:AB3")
            self.assertEqual(ingested["learned_layout"]["id_count"], 2)

            second_plan = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="students"))
            self.assertEqual([item["student"] for item in second_plan["cached"]], ["张三"])
            self.assertEqual([item["student"] for item in second_plan["read"]], ["李四"])
            self.assertEqual(second_plan["read"][0]["read_mode"], "learned_fast")
            self.assertEqual(second_plan["read"][0]["sheet"], "作业")
            self.assertEqual(second_plan["read"][0]["range"], "A1:AB3")

    def test_learned_range_mismatch_falls_back_only_for_that_student(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            homework = [
                {"student": "张三", "id": "student-zhang", "url": "https://docs.corp.kuaishou.com/student-zhang", "revision": "zhang-r1"},
                {"student": "李四", "id": "student-li", "url": "https://docs.corp.kuaishou.com/student-li", "revision": "li-r1"},
            ]
            payload = manifest_payload(homework=homework)
            manifest_path = temp / "task.json"
            write_json(manifest_path, payload)
            manifest = runtime.load_manifest(manifest_path)
            first = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="students"))
            probe = first["read"][0]
            probe_payload = document(["order", "画面美学"], [[1, "都一般"], [2, "一样好"]], student_name="张三", document_id="student-zhang", revision="zhang-r1")
            probe_payload.update({"sheet": "作业", "range": "A1:AB3"})
            probe_file = temp / "probe.json"
            write_json(probe_file, probe_payload)
            runtime.ingest_evidence(manifest, probe["item_id"], probe_file)

            second = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="students"))
            li = second["read"][0]
            mismatch_payload = document(["order", "画面美学"], [[1, "都一般"]], student_name="李四", document_id="student-li", revision="li-r1")
            mismatch_payload.update({"sheet": "例外作业", "range": "B2:AC3"})
            mismatch_file = temp / "mismatch.json"
            write_json(mismatch_file, mismatch_payload)
            result = runtime.ingest_evidence(manifest, li["item_id"], mismatch_file)
            self.assertEqual(result["status"], "needs_discovery")

            fallback = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="students"))
            fallback_li = next(item for item in fallback["read"] if item["student"] == "李四")
            self.assertEqual(fallback_li["read_mode"], "discovery_fallback")
            self.assertEqual(fallback_li["range"], "")
            fallback_payload = document(["order", "画面美学"], [[1, "都一般"], [2, "一样好"], [3, "一样差"]], student_name="李四", document_id="student-li", revision="li-r1")
            fallback_payload.update({"sheet": "例外作业", "range": "B2:AC5"})
            fallback_file = temp / "fallback.json"
            write_json(fallback_file, fallback_payload)
            completed = runtime.ingest_evidence(manifest, fallback_li["item_id"], fallback_file)
            self.assertEqual(completed["status"], "success")
            self.assertEqual(runtime.load_state(manifest)["layouts"]["homework"]["range"], "A1:AB3")
            final_plan = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="students"))
            cached_li = next(item for item in final_plan["cached"] if item["student"] == "李四")
            self.assertEqual(cached_li["read_mode"], "fixed_exception")
            self.assertEqual(cached_li["range"], "B2:AC5")

    def test_batch_ingest_and_retry_classification(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            manifest_path = temp / "task.json"
            write_json(manifest_path, manifest_payload())
            manifest = runtime.load_manifest(manifest_path)
            plan = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="students"))
            evidence_dir = temp / "incoming"
            evidence_dir.mkdir()
            for item in plan["read"]:
                payload = document(["ID", "画面美学"], [[1, "都一般"]], student_name=item["student"], document_id=item["document_id"], revision=item["revision"])
                write_json(evidence_dir / f"{item['item_id']}.json", payload)
            batch = runtime.ingest_evidence_batch(manifest, evidence_dir)
            self.assertEqual(batch["ingested"], 2)
            self.assertEqual(batch["failed"], 0)

            refreshed = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="students"), refresh=True)
            target = refreshed["read"][0]
            failure = runtime.record_failure(manifest, target["item_id"], "Docs timeout", "timeout")
            self.assertEqual(failure["status"], "pending_retry")
            waiting = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="students"), refresh=True)
            self.assertTrue(any(item["item_id"] == target["item_id"] for item in waiting["retry"]))

            fatal = runtime.record_failure(manifest, target["item_id"], "permission denied", "permission")
            self.assertEqual(fatal["status"], "failed")
            layout_failure = runtime.record_failure(manifest, target["item_id"], "已学习范围不再有效", "layout_mismatch")
            self.assertEqual(layout_failure["status"], "needs_discovery")
            fallback = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="students"))
            fallback_target = next(item for item in fallback["read"] if item["item_id"] == target["item_id"])
            self.assertEqual(fallback_target["read_mode"], "discovery_fallback")

    def test_preflight_reports_all_fatal_and_retryable_errors(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            manifest_path = temp / "task.json"
            write_json(manifest_path, manifest_payload())
            manifest = runtime.load_manifest(manifest_path)
            specs = runtime.document_specs(manifest, runtime.load_state(manifest), stage="students")
            snapshot = temp / "preflight.json"
            write_json(snapshot, {"items": [
                {"item_id": specs[0]["item_id"], "error_kind": "permission", "error": "无权访问"},
                {"item_id": specs[1]["item_id"], "error_kind": "timeout", "error": "Docs timeout"},
            ]})
            planned = runtime.plan_reads(manifest, runtime.merge_revisions(specs, snapshot))
            self.assertEqual([item["student"] for item in planned["failed"]], ["张三"])
            self.assertEqual([item["student"] for item in planned["retry"]], ["李四"])
            self.assertEqual(planned["read"], [])

            write_json(snapshot, {"items": [
                {"item_id": specs[0]["item_id"], "error_kind": "permission", "error": "无权访问"},
                {"item_id": specs[1]["item_id"], "revision": "li-r1"},
            ]})
            recovered = runtime.plan_reads(manifest, runtime.merge_revisions(specs, snapshot))
            self.assertEqual([item["student"] for item in recovered["read"]], ["李四"])
            self.assertEqual(recovered["read"][0]["read_mode"], "fixed")

    def test_preflight_retry_stops_at_configured_attempt_limit(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            manifest_path = temp / "task.json"
            payload = manifest_payload(homework=[{
                "student": "李四", "id": "student-li",
                "url": "https://docs.corp.kuaishou.com/student-li",
                "revision": "li-r1", "sheet": "Sheet1", "range": "A1:Z100",
                "evidence": "li.json",
            }])
            payload["runtime"] = {"retries": 3, "retry_delays_seconds": [1, 2]}
            write_json(manifest_path, payload)
            manifest = runtime.load_manifest(manifest_path)
            specs = runtime.document_specs(manifest, runtime.load_state(manifest), stage="students")
            snapshot = temp / "preflight.json"
            write_json(snapshot, {"items": [{
                "item_id": specs[0]["item_id"], "error_kind": "timeout", "error": "Docs timeout",
            }]})

            first = runtime.plan_reads(manifest, runtime.merge_revisions(specs, snapshot))
            second = runtime.plan_reads(manifest, runtime.merge_revisions(specs, snapshot))
            third = runtime.plan_reads(manifest, runtime.merge_revisions(specs, snapshot))

            self.assertEqual(first["retry"][0]["attempts"], 1)
            self.assertEqual(second["retry"][0]["attempts"], 2)
            self.assertEqual(third["retry"], [])
            self.assertEqual(third["failed"][0]["attempts"], 3)
            state = runtime.load_state(manifest)
            entry = state["documents"][specs[0]["item_id"]]
            self.assertEqual(entry["status"], "failed")
            self.assertEqual(entry["attempts"], 3)

    def test_index_resolves_links_and_removes_deleted_students(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            payload = manifest_payload()
            payload["homework"] = {
                "index": {
                    "id": "index-doc",
                    "url": "https://docs.corp.kuaishou.com/index-doc",
                    "revision": "index-r1",
                    "sheet": "索引",
                    "range": "A1:B100",
                    "name_header": "同学名称",
                    "link_header": "作业链接",
                }
            }
            manifest_path = temp / "task.json"
            write_json(manifest_path, payload)
            manifest = runtime.load_manifest(manifest_path)
            initial = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="initial"))
            index_item = next(item for item in initial["read"] if item["role"] == "homework_index")
            index_file = temp / "index.json"
            index_payload = document(
                ["同学名称", "作业链接"],
                [["张三", "https://docs.corp.kuaishou.com/zhang"], ["李四", "https://docs.corp.kuaishou.com/li"]],
                document_id="index-doc",
                revision="index-r1",
            )
            index_payload["sheet"] = "索引"
            index_payload["range"] = "A1:B100"
            write_json(index_file, index_payload)
            runtime.ingest_evidence(manifest, index_item["item_id"], index_file)
            students = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="students"))
            self.assertEqual(students["max_concurrency"], 15)
            self.assertEqual({item["student"] for item in students["read"] + students["deferred"]}, {"张三", "李四"})

            li_spec = next(item for item in students["read"] + students["deferred"] if item["student"] == "李四")
            runtime.record_failure(manifest, li_spec["item_id"], "模拟失败")
            index_payload["rows"] = [["张三", "https://docs.corp.kuaishou.com/zhang"]]
            write_json(index_file, index_payload)
            runtime.ingest_evidence(manifest, index_item["item_id"], index_file)
            state = runtime.load_state(manifest)
            self.assertEqual(state["documents"][li_spec["item_id"]]["status"], "removed")
            self.assertFalse(any(item.get("student") == "李四" for item in pipeline.manifest_failures(manifest)))

    def test_cached_index_restores_students_after_state_rebuild(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            payload = manifest_payload()
            payload["homework"] = {"index": {
                "id": "index-doc", "url": "https://docs.corp.kuaishou.com/index-doc", "revision": "index-r1",
                "sheet": "索引", "range": "A1:B3", "name_header": "同学名称", "link_header": "作业链接",
            }}
            manifest_path = temp / "task.json"
            write_json(manifest_path, payload)
            manifest = runtime.load_manifest(manifest_path)
            initial = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="initial"))
            index_item = next(item for item in initial["read"] if item["role"] == "homework_index")
            index_file = temp / "index.json"
            index_payload = document(
                ["同学名称", "作业链接"],
                [["张三", "https://docs.corp.kuaishou.com/zhang"]],
                document_id="index-doc", revision="index-r1",
            )
            index_payload.update({"sheet": "索引", "range": "A1:B3"})
            write_json(index_file, index_payload)
            runtime.ingest_evidence(manifest, index_item["item_id"], index_file)
            runtime.state_path(manifest).unlink()
            rebuilt = runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="initial"))
            self.assertEqual([item["role"] for item in rebuilt["cached"]], ["homework_index"])
            self.assertEqual(runtime.load_state(manifest)["resolved_students"][0]["student"], "张三")

    def test_student_identity_is_part_of_evidence_cache_key(self):
        base = {"document_id": "shared", "revision": "r1", "sheet": "作业", "range": "A1:C3", "role": "homework"}
        self.assertNotEqual(
            runtime.evidence_cache_key({**base, "student": "张三"}),
            runtime.evidence_cache_key({**base, "student": "李四"}),
        )

    def test_removed_explicit_document_is_not_left_active(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            manifest_path = temp / "task.json"
            payload = manifest_payload()
            write_json(manifest_path, payload)
            manifest = runtime.load_manifest(manifest_path)
            first_specs = runtime.document_specs(manifest, runtime.load_state(manifest), stage="students")
            runtime.plan_reads(manifest, first_specs)
            li_id = next(item["item_id"] for item in first_specs if item["student"] == "李四")
            payload["homework"]["documents"] = payload["homework"]["documents"][:1]
            write_json(manifest_path, payload)
            manifest = runtime.load_manifest(manifest_path)
            runtime.plan_reads(manifest, runtime.document_specs(manifest, runtime.load_state(manifest), stage="students"))
            self.assertEqual(runtime.load_state(manifest)["documents"][li_id]["status"], "removed")

    def test_manifest_incremental_scoring_and_json_default(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            manifest_path = temp / "task.json"
            write_json(manifest_path, manifest_payload())
            write_json(temp / "standard.json", document(["ID", "画面美学"], [[1, "都一般"], [2, "videoExp"]], document_id="standard-doc", revision="std-r1"))
            write_json(temp / "zhang.json", document(["ID", "画面美学"], [[1, "都一般"], [2, "videoExp好"]], student_name="张三", document_id="student-zhang", revision="zhang-r1"))
            write_json(temp / "li.json", document(["ID", "画面美学"], [[1, "一样差"], [2, "videoExp好"]], student_name="李四", document_id="student-li", revision="li-r1"))
            args = pipeline.parse_args(["--manifest", str(manifest_path), "--output", str(temp / "output")])

            xlsx, png, first = pipeline.run(args)
            self.assertIsNone(xlsx)
            self.assertIsNone(png)
            self.assertEqual(first["cache_stats"]["scores"], {"hits": 0, "misses": 2})
            self.assertTrue(Path(first["_result_path"]).exists())
            _, _, second = pipeline.run(args)
            self.assertEqual(second["cache_stats"]["scores"], {"hits": 2, "misses": 0})

            write_json(temp / "li.json", document(["ID", "画面美学"], [[1, "都一般"], [2, "videoExp好"]], student_name="李四", document_id="student-li", revision="li-r1"))
            _, _, changed_one = pipeline.run(args)
            self.assertEqual(changed_one["cache_stats"]["scores"], {"hits": 1, "misses": 1})

            write_json(temp / "standard.json", document(["ID", "画面美学"], [[1, "都一般"], [2, "videoContrast"]], document_id="standard-doc", revision="std-r1"))
            _, _, changed_standard = pipeline.run(args)
            self.assertEqual(changed_standard["cache_stats"]["scores"], {"hits": 0, "misses": 2})

    def test_manifest_standard_blank_review_keeps_homework_evidence_and_scores_only_after_fix(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            manifest_path = temp / "task.json"
            write_json(manifest_path, manifest_payload())
            write_json(temp / "standard.json", document(["ID", "画面美学"], [[1, ""]], document_id="standard-doc", revision="std-r1"))
            write_json(temp / "zhang.json", document(["ID", "画面美学"], [[1, "都一般"]], student_name="张三", document_id="student-zhang", revision="zhang-r1"))
            write_json(temp / "li.json", document(["ID", "画面美学"], [[1, "都一般"]], student_name="李四", document_id="student-li", revision="li-r1"))
            args = pipeline.parse_args(["--manifest", str(manifest_path), "--output", str(temp / "output")])

            _, _, waiting = pipeline.run(args)
            self.assertEqual(waiting["run_status"], "awaiting_standard_decision")
            self.assertEqual(waiting["cache_stats"]["scores"], {"hits": 0, "misses": 0})
            manifest = runtime.load_manifest(manifest_path)
            before = {
                item_id: (entry.get("content_sha256"), entry.get("cache_path"))
                for item_id, entry in runtime.load_state(manifest)["documents"].items()
                if entry.get("role") == "homework"
            }
            self.assertEqual(len(before), 2)

            write_json(temp / "standard.json", document(["ID", "画面美学"], [[1, "都一般"]], document_id="standard-doc", revision="std-r1"))
            _, _, completed = pipeline.run(args)
            self.assertEqual(completed["run_status"], "complete")
            after = {
                item_id: (entry.get("content_sha256"), entry.get("cache_path"))
                for item_id, entry in runtime.load_state(manifest)["documents"].items()
                if entry.get("role") == "homework"
            }
            self.assertEqual(after, before)
            self.assertEqual(completed["cache_stats"]["scores"], {"hits": 0, "misses": 2})

    def test_incomplete_manifest_pauses_outputs_and_resumes_only_missing_student(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            payload = manifest_payload()
            payload["output"] = {"dir": "output", "png": "on", "xlsx": "on"}
            manifest_path = temp / "task.json"
            write_json(manifest_path, payload)
            write_json(temp / "standard.json", document(["ID", "画面美学"], [[1, "都一般"]], document_id="standard-doc", revision="std-r1"))
            write_json(temp / "zhang.json", document(["ID", "画面美学"], [[1, "都一般"]], student_name="张三", document_id="student-zhang", revision="zhang-r1"))
            args = pipeline.parse_args(["--manifest", str(manifest_path), "--output", str(temp / "output")])
            xlsx, png, incomplete = pipeline.run(args)
            self.assertIsNone(xlsx)
            self.assertIsNone(png)
            self.assertEqual(incomplete["run_status"], "incomplete")
            self.assertEqual([item["student"] for item in incomplete["failed_documents"]], ["李四"])
            self.assertTrue(Path(incomplete["_result_path"]).exists())

            write_json(temp / "li.json", document(["ID", "画面美学"], [[1, "都一般"]], student_name="李四", document_id="student-li", revision="li-r1"))
            xlsx, png, complete = pipeline.run(args)
            self.assertEqual(complete["run_status"], "complete")
            self.assertTrue(xlsx and xlsx.exists())
            self.assertTrue(png and png.exists())
            self.assertEqual(complete["cache_stats"]["scores"], {"hits": 1, "misses": 1})

    def test_summary_only_saves_json_and_result_json_renders_outputs(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            standard = temp / "standard.json"
            homework = temp / "homework.json"
            output = temp / "output"
            write_json(standard, document(["ID", "画面美学"], [[1, "都一般"]]))
            write_json(homework, document(["同学名称", "ID", "画面美学"], [["王五", 1, "都一般"]]))
            args = pipeline.parse_args([
                "--exam-profile", "retake-image-aesthetic", "--group", "29组", "--progress", "补考画面美学",
                "--standard-evidence", str(standard), "--homework-evidence", str(homework), "--output", str(output),
                "--summary-only", "--png", "on", "--xlsx", "on",
            ])
            xlsx, png, result = pipeline.run(args)
            self.assertIsNone(xlsx)
            self.assertIsNone(png)
            result_path = Path(result["_result_path"])
            self.assertTrue(result_path.exists())

            render_dir = temp / "rendered"
            render_args = pipeline.parse_args([
                "--result-json", str(result_path), "--output", str(render_dir), "--png", "on", "--xlsx", "on",
            ])
            xlsx, png, rendered = pipeline.run(render_args)
            self.assertTrue(xlsx and xlsx.exists())
            self.assertTrue(png and png.exists())
            self.assertEqual(rendered["summary"], result["summary"])

    def test_result_json_rejects_incomplete_schema_cleanly(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            result_path = temp / "broken.json"
            write_json(result_path, {"schema_version": 2, "metadata": {"group": "29组"}})
            args = pipeline.parse_args(["--result-json", str(result_path), "--output", str(temp / "delivery")])
            with self.assertRaisesRegex(pipeline.AssessmentError, "缺少 run_status"):
                pipeline.run(args)

    def test_result_json_rejects_cross_field_contradictions(self):
        result = pipeline.build_scored_result(
            profile_key="retake-image-aesthetic", group="29组", progress="语义门禁",
            standard_documents=[document(["ID", "画面美学"], [[1, "都一般"]])],
            homework_documents=[document(["同学名称", "ID", "画面美学"], [["王五", 1, "都一般"]])],
            aliases={}, source_mode="docs",
        )
        contradictions = []
        failed_complete = json.loads(json.dumps(result, ensure_ascii=False))
        failed_complete["failed_documents"] = [{"role": "homework", "error": "读取失败"}]
        contradictions.append((failed_complete, "failed_documents"))
        bad_math = json.loads(json.dumps(result, ensure_ascii=False))
        bad_math["summary"][0]["dimensions"]["画面美学"]["accuracy"] = 1.5
        contradictions.append((bad_math, "accuracy"))
        bad_details = json.loads(json.dumps(result, ensure_ascii=False))
        bad_details["details"][0]["result"] = "错误"
        contradictions.append((bad_details, "summary与逐题明细"))
        empty = json.loads(json.dumps(result, ensure_ascii=False))
        empty["summary"] = []
        empty["details"] = []
        contradictions.append((empty, "至少包含一名学员"))
        for payload, message in contradictions:
            with self.subTest(message=message), self.assertRaisesRegex(pipeline.AssessmentError, message):
                pipeline.validate_result_schema(payload)

    def test_each_run_is_atomically_isolated_and_never_reuses_old_artifact_names(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            output = temp / "delivery"
            complete = pipeline.build_scored_result(
                profile_key="retake-image-aesthetic", group="29组", progress="生命周期",
                standard_documents=[document(["ID", "画面美学"], [[1, "都一般"]])],
                homework_documents=[document(["同学名称", "ID", "画面美学"], [["王五", 1, "都一般"]])],
                aliases={}, source_mode="docs",
            )
            xlsx, png, first, first_json, _ = pipeline.deliver_result(complete, output, "on", "on")
            self.assertTrue(xlsx and xlsx.exists())
            self.assertTrue(png and png.exists())
            incomplete = json.loads(json.dumps(complete, ensure_ascii=False))
            incomplete["run_status"] = "incomplete"
            incomplete["failed_documents"] = [{"role": "homework", "student": "李四", "error": "无权访问"}]
            incomplete["stopped_items"] = pipeline.stopped_items_for(incomplete)
            _, _, second, second_json, _ = pipeline.deliver_result(incomplete, output, "on", "on")
            self.assertNotEqual(first_json.parent, second_json.parent)
            self.assertTrue(first_json.exists() and second_json.exists())
            self.assertIsNone(second["outputs"]["png"])
            self.assertIsNone(second["outputs"]["xlsx"])
            self.assertEqual(list(second_json.parent.glob("*.png")), [])
            self.assertEqual(list(second_json.parent.glob("*.xlsx")), [])

    def test_output_failure_rolls_back_all_formal_artifacts_and_reports_reason(self):
        with tempfile.TemporaryDirectory() as temp_name:
            result = pipeline.build_scored_result(
                profile_key="retake-image-aesthetic", group="29组", progress="原子回滚",
                standard_documents=[document(["ID", "画面美学"], [[1, "都一般"]])],
                homework_documents=[document(["同学名称", "ID", "画面美学"], [["王五", 1, "都一般"]])],
                aliases={}, source_mode="docs",
            )
            with patch.object(pipeline, "build_workbook", side_effect=pipeline.AssessmentError("模拟Excel失败")):
                xlsx, png, failed, result_path, hard = pipeline.deliver_result(result, Path(temp_name), "on", "on")
            self.assertTrue(hard)
            self.assertEqual(failed["run_status"], "output_failed")
            self.assertIsNone(xlsx)
            self.assertIsNone(png)
            self.assertTrue(result_path.exists())
            self.assertEqual(list(result_path.parent.glob("*.png")), [])
            self.assertEqual(list(result_path.parent.glob("*.xlsx")), [])
            self.assertTrue(any(item["stage"] == "output" and "模拟Excel失败" in item["reason"] for item in failed["stopped_items"]))
            retry_args = pipeline.parse_args([
                "--result-json", str(result_path), "--output", str(Path(temp_name) / "retry"),
                "--png", "on", "--xlsx", "on",
            ])
            retry_xlsx, retry_png, recovered = pipeline.run(retry_args)
            self.assertEqual(recovered["run_status"], "complete")
            self.assertTrue(retry_xlsx and retry_xlsx.exists())
            self.assertTrue(retry_png and retry_png.exists())

    def test_direct_evidence_uses_same_strict_validator_as_manifest_ingest(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            missing_id = document(["ID", "画面美学"], [[1, "都一般"]])
            missing_id["document"]["id"] = ""
            missing_id_path = temp / "missing-id.json"
            write_json(missing_id_path, missing_id)
            with self.assertRaisesRegex(pipeline.AssessmentError, "document.id"):
                pipeline.load_evidence_files([missing_id_path])
            object_row = document(["ID", "画面美学"], [{"ID": 1, "画面美学": "都一般"}])
            object_row_path = temp / "object-row.json"
            write_json(object_row_path, object_row)
            with self.assertRaisesRegex(pipeline.AssessmentError, "必须是数组"):
                pipeline.load_evidence_files([object_row_path])

    def test_output_path_that_is_a_file_is_rejected_cleanly(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            output_file = temp / "not-a-directory"
            output_file.write_text("x", encoding="utf-8")
            result_path = temp / "result.json"
            payload = pipeline.build_scored_result(
                profile_key="retake-image-aesthetic", group="29组", progress="测试",
                standard_documents=[document(["ID", "画面美学"], [[1, "都一般"]])],
                homework_documents=[document(["同学名称", "ID", "画面美学"], [["王五", 1, "都一般"]])],
                aliases={}, source_mode="docs",
            )
            write_json(result_path, payload)
            args = pipeline.parse_args(["--result-json", str(result_path), "--output", str(output_file)])
            with self.assertRaisesRegex(pipeline.AssessmentError, "文件而不是文件夹"):
                pipeline.run(args)

    def test_explicit_relative_output_path_is_rejected(self):
        args = pipeline.parse_args(["--result-json", "/tmp/result.json", "--output", "relative-delivery"])
        with self.assertRaisesRegex(pipeline.AssessmentError, "绝对交付目录"):
            pipeline.output_directory(args)

    def test_manifest_cli_reports_clean_error_without_traceback(self):
        with tempfile.TemporaryDirectory() as temp_name:
            missing = Path(temp_name) / "missing.json"
            completed = subprocess.run(
                [sys.executable, str(Path(runtime.__file__)), "status", "--manifest", str(missing)],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(completed.returncode, 2)
            self.assertIn("错误：", completed.stderr)
            self.assertNotIn("Traceback", completed.stderr)
            stopped = json.loads(completed.stderr.strip().splitlines()[-1])
            self.assertTrue(stopped["stopped_items"][0]["reason"])
            self.assertTrue(stopped["stopped_items"][0]["next_action"])
            invalid_args = subprocess.run(
                [sys.executable, str(Path(runtime.__file__)), "--not-a-real-option"],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(invalid_args.returncode, 2)
            invalid_payload = json.loads(invalid_args.stderr.strip().splitlines()[-1])
            self.assertEqual(invalid_payload["status"], "stopped")
            self.assertIn("命令参数错误", invalid_payload["stopped_items"][0]["reason"])
            self.assertTrue(invalid_payload["stopped_items"][0]["next_action"])

    def test_assessment_cli_always_reports_why_execution_stopped(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            completed = subprocess.run(
                [
                    sys.executable, str(MODULE_PATH), "--result-json", str(temp / "missing.json"),
                    "--output", str(temp / "delivery"),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(completed.returncode, 2)
            payload = json.loads(completed.stderr.strip().splitlines()[-1])
            self.assertEqual(payload["status"], "stopped")
            self.assertTrue(payload["stopped_items"][0]["reason"])
            self.assertTrue(payload["stopped_items"][0]["next_action"])
            invalid_args = subprocess.run(
                [sys.executable, str(MODULE_PATH), "--png", "invalid"],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(invalid_args.returncode, 2)
            invalid_payload = json.loads(invalid_args.stderr.strip().splitlines()[-1])
            self.assertIn("命令参数错误", invalid_payload["stopped_items"][0]["reason"])
            unknown_args = subprocess.run(
                [sys.executable, str(MODULE_PATH), "--not-a-real-option"],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(unknown_args.returncode, 2)
            unknown_payload = json.loads(unknown_args.stderr.strip().splitlines()[-1])
            self.assertEqual(unknown_payload["status"], "stopped")
            self.assertIn("命令参数错误", unknown_payload["stopped_items"][0]["reason"])
            self.assertTrue(unknown_payload["stopped_items"][0]["next_action"])

    def test_flat_layout_package_uses_skill_name_for_file_and_archive_root(self):
        with tempfile.TemporaryDirectory() as temp_name:
            flat = Path(temp_name)
            script_location = Path(__file__).resolve().parent
            project_root = script_location.parent if (script_location.parent / "SKILL.md").is_file() else script_location
            standard_layout = (project_root / "scripts" / "run_assessment.py").is_file()

            def source(standard_name, flat_name):
                return project_root / (standard_name if standard_layout else flat_name)

            shutil.copy2(project_root / "SKILL.md", flat / "SKILL.md")
            shutil.copy2(project_root / "requirements.txt", flat / "requirements.txt")
            shutil.copy2(source("agents/openai.yaml", "openai.yaml"), flat / "openai.yaml")
            for name in ("exam_profiles.json", "evidence-schema.md", "manifest.md"):
                shutil.copy2(source(f"references/{name}", name), flat / name)
            for name in (
                "run_assessment.py", "manifest_runtime.py", "build_workbook.py", "ocr_api.py",
                "ocr_vision.swift", "package_skill.py", "test_pipeline.py",
            ):
                shutil.copy2(source(f"scripts/{name}", name), flat / name)
            completed = subprocess.run(
                [sys.executable, str(flat / "package_skill.py")],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(completed.returncode, 0, completed.stderr)
            archive_path = flat / "dist" / "score-kuaishou-exams.zip"
            self.assertTrue(archive_path.is_file())
            with zipfile.ZipFile(archive_path) as archive:
                names = archive.namelist()
            self.assertTrue(names)
            self.assertTrue(all(name.startswith("score-kuaishou-exams/") for name in names))

    def test_zero_student_result_is_incomplete(self):
        result = pipeline.build_scored_result(
            profile_key="retake-image-aesthetic", group="29组", progress="空作业",
            standard_documents=[document(["ID", "画面美学"], [[1, "都一般"]])],
            homework_documents=[], aliases={}, source_mode="docs",
        )
        self.assertEqual(result["run_status"], "incomplete")
        self.assertIn("没有发现任何有效学员", result["failed_documents"][0]["error"])

    def test_excel_escapes_formula_like_user_text(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            standard = temp / "standard.json"
            homework = temp / "homework.json"
            output = temp / "delivery"
            write_json(standard, document(["ID", "画面美学"], [[1, "都一般"]]))
            write_json(homework, document(["同学名称", "ID", "画面美学"], [["=1+1", 1, "都一般"]]))
            args = pipeline.parse_args([
                "--exam-profile", "retake-image-aesthetic", "--group", "29组", "--progress", "公式防护",
                "--standard-evidence", str(standard), "--homework-evidence", str(homework),
                "--output", str(output), "--xlsx", "on",
            ])
            xlsx, _, _ = pipeline.run(args)
            workbook = load_workbook(xlsx, data_only=False)
            self.assertEqual(workbook["成绩汇总"]["C2"].value, "'=1+1")
            workbook.close()

    def test_excel_sanitizes_leading_whitespace_formulas_and_illegal_xml_controls(self):
        import build_workbook as workbook_builder

        self.assertEqual(workbook_builder.safe_excel_value(" \t=1+1"), "' \t=1+1")
        self.assertEqual(workbook_builder.safe_excel_value("\r@cmd"), "'\r@cmd")
        self.assertEqual(workbook_builder.safe_excel_value("王\x00五"), "王�五")

    def test_end_to_end_png_without_workbook(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            standard = temp / "standard.json"
            homework = temp / "homework.json"
            output = temp / "output"
            standard.write_text(json.dumps(document(["ID", "画面美学"], [[1, "般"], [2, "一样好"]]), ensure_ascii=False), encoding="utf-8")
            homework.write_text(json.dumps(document(["同学名称", "ID", "画面美学"], [["王五", 2, "一样差"], ["王五", 1, "般"]]), ensure_ascii=False), encoding="utf-8")
            args = pipeline.parse_args([
                "--exam-profile", "retake-image-aesthetic",
                "--group", "29组",
                "--progress", "补考画面美学",
                "--standard-evidence", str(standard),
                "--homework-evidence", str(homework),
                "--output", str(output),
                "--skip-xlsx",
                "--png", "on",
            ])
            _, png, result = pipeline.run(args)
            self.assertTrue(png.exists())
            self.assertGreater(png.stat().st_size, 1000)
            self.assertEqual(result["summary"][0]["dimensions"]["画面美学"]["accuracy"], 0.5)

    def test_end_to_end_xlsx_and_png(self):
        with tempfile.TemporaryDirectory() as temp_name:
            temp = Path(temp_name)
            standard = temp / "standard.json"
            homework = temp / "homework.json"
            output = temp / "output"
            standard.write_text(
                json.dumps(document(["ID", "画面美学", "动态美学"], [[1, ["都一般", "一样好"], "一样好"], [2, "videoExp", "videoContrast好"]]), ensure_ascii=False),
                encoding="utf-8",
            )
            homework.write_text(
                json.dumps(
                    document(
                        ["同学名称", "动态美学", "ID", "画面美学", "解释"],
                        [
                            ["王五", "videoContrast好", 2, "videoExp好", "忽略"],
                            ["王五", "一样好", 1, "答案：都一般", "忽略"],
                            ["赵六", "一样差", 1, "一样好", "忽略"],
                            ["赵六", "videoContrast", 2, "videoExp好", "忽略"],
                        ],
                    ),
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            args = pipeline.parse_args([
                "--exam-profile", "online-aesthetics-dual",
                "--group", "29组",
                "--progress", "画面美学＋动态美学",
                "--standard-evidence", str(standard),
                "--homework-evidence", str(homework),
                "--output", str(output),
                "--png", "on",
                "--xlsx", "on",
            ])
            xlsx, png, result = pipeline.run(args)
            self.assertIsNotNone(xlsx)
            self.assertTrue(xlsx.exists())
            self.assertGreater(xlsx.stat().st_size, 5000)
            self.assertTrue(png.exists())
            self.assertEqual(len(result["summary"]), 2)
            multi_details = [item for item in result["details"] if item["dimension"] == "画面美学" and item["id"] == "1"]
            self.assertTrue(all(item["result"] == "正确" for item in multi_details))
            self.assertTrue(all(item["standard_keywords"] == "都一般｜一样好" for item in multi_details))
            evidence = [item for item in result["standard_answer_evidence"] if item["dimension"] == "画面美学" and item["id"] == "1"]
            self.assertEqual(evidence[0]["raw_cell"], ["都一般", "一样好"])
            self.assertEqual(evidence[0]["keywords"], ["都一般", "一样好"])
            reverse_match = [item for item in result["details"] if item["student"] == "赵六" and item["dimension"] == "动态美学" and item["id"] == "2"]
            self.assertEqual(reverse_match[0]["result"], "正确")
            self.assertEqual(reverse_match[0]["matched_keywords"], "videoContrast好")
            workbook = load_workbook(xlsx, data_only=False)
            self.assertEqual(workbook.sheetnames, ["成绩汇总", "逐题明细", "异常复核", "证据索引"])
            self.assertIsInstance(workbook["成绩汇总"]["D2"].value, (int, float))
            self.assertGreater(len(workbook["成绩汇总"].conditional_formatting), 0)
            workbook.close()


if __name__ == "__main__":
    unittest.main(verbosity=2)
