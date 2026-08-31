#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import re
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="D8E5F7")
HEADER_FONT = Font(name="PingFang SC", size=10, bold=True, color="25324B")
BODY_FONT = Font(name="PingFang SC", size=10, color="344054")
THIN = Side(style="thin", color="E4E7EC")
HEADER_SIDE = Side(style="thin", color="C5D3E7")
SCORE_COLORS = [
    ("greaterThanOrEqual", ["0.9"], "6AB37B", "25324B"),
    ("between", ["0.8", "0.899999999"], "A7D08D", "25324B"),
    ("between", ["0.7", "0.799999999"], "FEE07B", "25324B"),
    ("between", ["0.6", "0.699999999"], "F5A05C", "25324B"),
    ("lessThan", ["0.6"], "F35161", "FFFFFF"),
]
ILLEGAL_XML_CONTROL = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def safe_excel_value(value: Any) -> Any:
    """Prevent user-controlled text from becoming an Excel formula."""
    if isinstance(value, str):
        value = ILLEGAL_XML_CONTROL.sub("�", value)
        if value.lstrip().startswith(("=", "+", "-", "@")):
            return "'" + value
    return value


def append_safe(ws, values: list[Any]) -> None:
    ws.append([safe_excel_value(value) for value in values])


def style_sheet(ws, widths: list[float], *, percent_columns: list[int] | None = None) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 34
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=HEADER_SIDE, right=HEADER_SIDE, top=HEADER_SIDE, bottom=HEADER_SIDE)
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 30
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for column in percent_columns or []:
        for cells in ws.iter_cols(min_col=column, max_col=column, min_row=2, max_row=max(2, ws.max_row)):
            for cell in cells:
                cell.number_format = "0.00%"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape" if ws.max_column > 5 else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3 if ws.max_column > 8 else ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:1"
    ws.print_area = ws.dimensions
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4


def add_score_rules(ws, cell_range: str) -> None:
    for operator, formulas, fill_color, font_color in SCORE_COLORS:
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator=operator,
                formula=formulas,
                fill=PatternFill("solid", fgColor=fill_color),
                font=Font(color=font_color, bold=fill_color in {"6AB37B", "F35161"}),
            ),
        )


def accuracy_cell(rate: Any, status: str) -> Any:
    if rate is not None:
        return rate
    return "未评分" if status == "未评分" else "待复核"


def build_summary(ws, data: dict[str, Any]) -> None:
    dimensions = list(data["metadata"]["dimensions"])
    summaries = list(data.get("summary", []))
    if len(dimensions) == 1:
        dimension = dimensions[0]
        summaries.sort(
            key=lambda item: (
                item["dimensions"][dimension]["accuracy"] is None,
                -(item["dimensions"][dimension]["accuracy"] or -1),
                item.get("source_order", 0),
            )
        )
        append_safe(ws, ["组别", "进度", "同学", "答对数", f"{dimension}准确率", "错误 ID"])
        for summary in summaries:
            dim = summary["dimensions"][dimension]
            wrong = list(dim.get("wrong_ids", [])) + [f"{item}*" for item in dim.get("review_ids", [])] + [f"{item}(排除)" for item in dim.get("excluded_ids", [])]
            append_safe(
                ws,
                [
                    data["metadata"]["group"],
                    data["metadata"]["progress"],
                    summary["student"],
                    "未评分" if dim.get("status") == "未评分" else (f"{dim['correct']}/{dim['total']}" if dim.get("accuracy") is not None else "待复核"),
                    accuracy_cell(dim.get("accuracy"), dim.get("status", "")),
                    ", ".join(wrong) or "—",
                ]
            )
        style_sheet(ws, [12, 18, 18, 14, 18, 34], percent_columns=[5])
        if ws.max_row >= 2:
            add_score_rules(ws, f"E2:E{ws.max_row}")
        return

    append_safe(ws, ["组别", "进度", "姓名", *dimensions, "总准确率（有效格汇总）"])
    for summary in summaries:
        append_safe(
            ws,
            [
                data["metadata"]["group"],
                data["metadata"]["progress"],
                summary["student"],
                *[
                    accuracy_cell(
                        summary["dimensions"][dimension].get("accuracy"),
                        summary["dimensions"][dimension].get("status", ""),
                    )
                    for dimension in dimensions
                ],
                accuracy_cell(summary.get("overall_accuracy"), summary.get("status", "")),
            ]
        )
    first_score = 4
    last_score = 4 + len(dimensions)
    style_sheet(ws, [12, 18, 18, *[20 for _ in dimensions], 22], percent_columns=list(range(first_score, last_score + 1)))
    if ws.max_row >= 2:
        add_score_rules(ws, f"D2:{get_column_letter(last_score)}{ws.max_row}")


def build_details(ws, data: dict[str, Any]) -> None:
    append_safe(ws, ["同学", "维度", "ID", "标准答案原文", "作业答案原文", "标准关键词", "命中关键词", "结果", "来源", "标准答案OCR置信度", "作业OCR置信度", "备注"])
    for item in data.get("details", []):
        append_safe(
            ws,
            [
                item.get("student", ""),
                item.get("dimension", ""),
                item.get("id", ""),
                item.get("standard_raw", ""),
                item.get("homework_raw", ""),
                item.get("standard_keywords", item.get("standard_canonical", "")),
                item.get("matched_keywords", item.get("homework_canonical", "")),
                item.get("result", ""),
                item.get("source", ""),
                item.get("standard_confidence", ""),
                item.get("confidence", ""),
                item.get("note", ""),
            ]
        )
    style_sheet(ws, [16, 24, 10, 24, 24, 20, 20, 12, 14, 16, 16, 42])
    for column in ("J", "K"):
        for cell in ws[column][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000"


def build_anomalies(ws, data: dict[str, Any]) -> None:
    append_safe(ws, ["同学", "异常类型", "维度", "ID", "说明"])
    rows = data.get("anomalies", [])
    if rows:
        for item in rows:
            append_safe(ws, [item.get("student", ""), item.get("type", ""), item.get("dimension", ""), item.get("id", ""), item.get("detail", "")])
    else:
        append_safe(ws, ["", "无异常", "", "", "本次运行未发现异常"])
    for failed in data.get("failed_documents", []):
        append_safe(ws, [failed.get("student", ""), "文档读取失败", "", "", failed.get("error", "")])
    style_sheet(ws, [18, 18, 24, 12, 52])


def build_evidence(ws, data: dict[str, Any]) -> None:
    append_safe(ws, ["来源", "本地证据路径", "文档链接", "文档ID", "Revision", "工作表/图片", "范围", "维度", "维度列绑定", "读取时间", "内容SHA-256"])
    for item in data.get("evidence", []):
        append_safe(
            ws,
            [
                item.get("source", ""),
                item.get("source_path", ""),
                item.get("url", ""),
                item.get("document_id", ""),
                item.get("revision", ""),
                item.get("sheet", ""),
                item.get("range", ""),
                "、".join(item.get("dimensions") or []),
                json.dumps(item.get("dimension_bindings") or {}, ensure_ascii=False, separators=(",", ":")),
                item.get("read_at", ""),
                item.get("content_sha256", ""),
            ]
        )
    style_sheet(ws, [14, 46, 52, 20, 18, 22, 14, 20, 46, 24, 48])
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 42


def validate_workbook(path: Path, data: dict[str, Any]) -> None:
    workbook = load_workbook(path, data_only=False)
    expected_sheets = ["成绩汇总", "逐题明细", "异常复核", "证据索引"]
    if workbook.sheetnames != expected_sheets:
        raise RuntimeError(f"Excel工作表不完整：{workbook.sheetnames}")
    if workbook["成绩汇总"].max_row != len(data.get("summary", [])) + 1:
        raise RuntimeError("Excel汇总人数与评分JSON不一致")
    if workbook["逐题明细"].max_row != len(data.get("details", [])) + 1:
        raise RuntimeError("Excel逐题明细数量与评分JSON不一致")
    if data.get("summary") and len(workbook["成绩汇总"].conditional_formatting) == 0:
        raise RuntimeError("Excel成绩汇总缺少原生条件格式")
    workbook.close()


def build_workbook(data: dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    build_summary(workbook.create_sheet("成绩汇总"), data)
    build_details(workbook.create_sheet("逐题明细"), data)
    build_anomalies(workbook.create_sheet("异常复核"), data)
    build_evidence(workbook.create_sheet("证据索引"), data)
    with tempfile.NamedTemporaryFile(prefix="score-kuaishou-", suffix=".xlsx", dir=output_path.parent, delete=False) as handle:
        temporary = Path(handle.name)
    try:
        workbook.save(temporary)
        validate_workbook(temporary, data)
        os.replace(temporary, output_path)
    finally:
        temporary.unlink(missing_ok=True)
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description="使用openpyxl生成快手考试Excel")
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    data = json.loads(args.input.read_text(encoding="utf-8"))
    build_workbook(data, args.output)
    print(json.dumps({"xlsx": str(args.output)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
